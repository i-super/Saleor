import shutil
from tempfile import NamedTemporaryFile
from unittest.mock import ANY, MagicMock, patch

import openpyxl
import petl as etl
import pytest
from django.core.files import File
from freezegun import freeze_time

from ....core import JobStatus
from ....graphql.csv.enums import ProductFieldEnum
from ....product.models import Product
from ... import FileTypes
from ...models import ExportFile
from ...utils.export import (
    append_to_file,
    create_file_with_headers,
    export_products,
    export_products_in_batches,
    get_filename,
    get_product_queryset,
    save_csv_file_in_export_file,
)


@pytest.mark.skip(reason="CSV export rewritten, needs update")
@pytest.mark.parametrize(
    "file_type", [FileTypes.CSV, FileTypes.XLSX],
)
@patch("saleor.csv.utils.export.create_file_with_headers")
@patch("saleor.csv.utils.export.export_products_in_batches")
@patch("saleor.csv.utils.export.send_email_with_link_to_download_file")
def test_export_products(
    send_email_mock,
    export_products_in_batches_mock,
    create_file_with_headers_mock,
    product_list,
    user_export_file,
    file_type,
):
    # given
    export_info = {
        "fields": [ProductFieldEnum.NAME.value],
        "warehouses": [],
        "attributes": [],
    }

    # when
    export_products(user_export_file, {"all": ""}, export_info, file_type)

    # then
    create_file_with_headers_mock.called_once_with(
        ["id", "name"], ";", user_export_file, ANY, file_type
    )
    export_products_in_batches_mock.called_once_with(
        Product.objects.all(),
        export_info,
        {"id", "name"},
        ["id", "name"],
        ";",
        user_export_file,
        file_type,
    )
    send_email_mock.called_once_with(
        user_export_file, user_export_file.user.email, "export_products_success"
    )


@pytest.mark.skip(reason="CSV export rewritten, needs update")
@patch("saleor.csv.utils.export.create_file_with_headers")
@patch("saleor.csv.utils.export.export_products_in_batches")
@patch("saleor.csv.utils.export.send_email_with_link_to_download_file")
def test_export_products_ids(
    send_email_mock,
    export_products_in_batches_mock,
    create_file_with_headers_mock,
    product_list,
    user_export_file,
):
    # given
    pks = [product.pk for product in product_list[:2]]
    export_info = {"fields": [], "warehouses": [], "attributes": []}
    file_type = FileTypes.CSV

    assert user_export_file.status == JobStatus.PENDING
    assert not user_export_file.content_file

    # when
    export_products(user_export_file, {"ids": pks}, export_info, file_type)

    # then
    create_file_with_headers_mock.called_once_with(
        ["id"], ";", user_export_file, ANY, file_type
    )
    export_products_in_batches_mock.called_once_with(
        Product.objects.filter(pk__in=pks),
        export_info,
        {"id"},
        ["id"],
        ";",
        user_export_file,
        file_type,
    )
    send_email_mock.called_once_with(
        user_export_file, user_export_file.user.email, "export_products_success"
    )


@pytest.mark.skip(reason="CSV export rewritten, needs update")
@patch("saleor.csv.utils.export.create_file_with_headers")
@patch("saleor.csv.utils.export.export_products_in_batches")
@patch("saleor.csv.utils.export.send_email_with_link_to_download_file")
def test_export_products_filter(
    send_email_mock,
    export_products_in_batches_mock,
    create_file_with_headers_mock,
    product_list,
    user_export_file,
):
    # given
    product_list[0].is_published = False
    product_list[0].save(update_fields=["is_published"])

    export_info = {"fields": [], "warehouses": [], "attributes": []}
    file_type = FileTypes.CSV

    assert user_export_file.status == JobStatus.PENDING
    assert not user_export_file.content_file

    # when
    export_products(
        user_export_file, {"filter": {"is_published": True}}, export_info, file_type
    )

    # then
    create_file_with_headers_mock.called_once_with(
        ["id"], ";", user_export_file, ANY, file_type
    )
    export_products_in_batches_mock.called_once_with(
        Product.objects.filter(is_published=True),
        export_info,
        {"id"},
        ["id"],
        ";",
        user_export_file,
        file_type,
    )
    send_email_mock.called_once_with(
        user_export_file, user_export_file.user.email, "export_products_success"
    )


@pytest.mark.skip(reason="CSV export rewritten, needs update")
@patch("saleor.csv.utils.export.create_file_with_headers")
@patch("saleor.csv.utils.export.export_products_in_batches")
@patch("saleor.csv.utils.export.send_email_with_link_to_download_file")
def test_export_products_by_app(
    send_email_mock,
    export_products_in_batches_mock,
    create_file_with_headers_mock,
    product_list,
    app_export_file,
):
    # given
    export_info = {
        "fields": [ProductFieldEnum.NAME.value],
        "warehouses": [],
        "attributes": [],
    }
    file_type = FileTypes.CSV

    # when
    export_products(app_export_file, {"all": ""}, export_info, file_type)

    # then
    create_file_with_headers_mock.called_once_with(
        ["id", "name"], ";", app_export_file, ANY, file_type
    )
    export_products_in_batches_mock.called_once_with(
        Product.objects.all(),
        export_info,
        {"id", "name"},
        ["id", "name"],
        ";",
        app_export_file,
        file_type,
    )
    send_email_mock.assert_not_called()


def test_get_filename_csv():
    with freeze_time("2000-02-09"):
        file_name = get_filename("test", FileTypes.CSV)

        assert file_name == "test_data_09_02_2000.csv"


def test_get_filename_xlsx():
    with freeze_time("2000-02-09"):
        file_name = get_filename("test", FileTypes.XLSX)

        assert file_name == "test_data_09_02_2000.xlsx"


def test_get_product_queryset_all(product_list):
    queryset = get_product_queryset({"all": ""})

    assert queryset.count() == len(product_list)


def test_get_product_queryset_ids(product_list):
    pks = [product.pk for product in product_list[:2]]
    queryset = get_product_queryset({"ids": pks})

    assert queryset.count() == len(pks)


def get_product_queryset_filter(product_list):
    product_not_published = product_list.first()
    product_not_published.is_published = False
    product_not_published.save()

    queryset = get_product_queryset({"ids": {"is_published": True}})

    assert queryset.count() == len(product_list) - 1


@pytest.mark.skip(reason="CSV export rewritten, needs update")
def test_create_file_with_headers_csv(user_export_file, tmpdir, media_root):
    # given
    file_headers = ["id", "name", "collections"]
    file_name = "test.csv"
    export_file_csv_upload_dir = ExportFile.content_file.field.upload_to

    assert not user_export_file.content_file

    # when
    create_file_with_headers(
        file_headers, ";", user_export_file, file_name, FileTypes.CSV
    )

    # then
    csv_file = user_export_file.content_file
    assert csv_file
    assert csv_file.name == f"{export_file_csv_upload_dir}/{file_name}"

    file_content = csv_file.read().decode().split("\r\n")

    assert ";".join(file_headers) in file_content

    shutil.rmtree(tmpdir)


@pytest.mark.skip(reason="CSV export rewritten, needs update")
def test_create_file_with_headers_xlsx(user_export_file, tmpdir, media_root):
    # given
    file_headers = ["id", "name", "collections"]
    file_name = "test.xlsx"
    export_file_csv_upload_dir = ExportFile.content_file.field.upload_to

    assert not user_export_file.content_file

    # when
    create_file_with_headers(
        file_headers, ";", user_export_file, file_name, FileTypes.XLSX
    )

    # then
    xlsx_file = user_export_file.content_file
    assert xlsx_file
    assert xlsx_file.name == f"{export_file_csv_upload_dir}/{file_name}"

    wb_obj = openpyxl.load_workbook(xlsx_file)

    sheet_obj = wb_obj.active
    max_col = sheet_obj.max_column
    headers = [sheet_obj.cell(row=1, column=i).value for i in range(1, max_col + 1)]

    assert headers == file_headers

    shutil.rmtree(tmpdir)


def test_save_csv_file_in_export_file(user_export_file, tmpdir, media_root):
    file_mock = MagicMock(spec=File)
    file_mock.name = "temp_file.csv"
    file_name = "test.csv"

    assert not user_export_file.content_file

    save_csv_file_in_export_file(user_export_file, file_mock, file_name)

    user_export_file.refresh_from_db()
    assert user_export_file.content_file

    shutil.rmtree(tmpdir)


@pytest.mark.skip(reason="CSV export rewritten, needs update")
def test_append_to_file_for_csv(user_export_file, tmpdir, media_root):
    # given
    export_data = [
        {"id": "123", "name": "test1", "collections": "coll1"},
        {"id": "345", "name": "test2"},
    ]
    headers = ["id", "name", "collections"]
    delimiter = ";"

    file_name = "test.csv"

    table = etl.fromdicts([{"id": "1", "name": "A"}], header=headers, missing=" ")

    with NamedTemporaryFile() as temp_file:
        etl.tocsv(table, temp_file.name, delimiter=delimiter)
        user_export_file.content_file.save(file_name, temp_file)

    # when
    append_to_file(export_data, headers, user_export_file, FileTypes.CSV, delimiter)

    # then
    user_export_file.refresh_from_db()

    csv_file = user_export_file.content_file
    file_content = csv_file.read().decode().split("\r\n")
    assert ";".join(headers) in file_content
    assert ";".join(export_data[0].values()) in file_content
    assert (";".join(export_data[1].values()) + "; ") in file_content

    shutil.rmtree(tmpdir)


@pytest.mark.skip(reason="CSV export rewritten, needs update")
def test_append_to_file_for_xlsx(user_export_file, tmpdir, media_root):
    # given
    export_data = [
        {"id": "123", "name": "test1", "collections": "coll1"},
        {"id": "345", "name": "test2"},
    ]
    expected_headers = ["id", "name", "collections"]
    delimiter = ";"

    file_name = "test.xlsx"

    table = etl.fromdicts(
        [{"id": "1", "name": "A"}], header=expected_headers, missing=" "
    )

    with NamedTemporaryFile() as temp_file:
        etl.io.xlsx.toxlsx(table, temp_file.name)
        user_export_file.content_file.save(file_name, temp_file)

    # when
    append_to_file(
        export_data, expected_headers, user_export_file, FileTypes.XLSX, delimiter
    )

    # then
    user_export_file.refresh_from_db()

    xlsx_file = user_export_file.content_file
    wb_obj = openpyxl.load_workbook(xlsx_file)

    sheet_obj = wb_obj.active
    max_col = sheet_obj.max_column
    max_row = sheet_obj.max_row
    expected_headers = expected_headers
    headers = [sheet_obj.cell(row=1, column=i).value for i in range(1, max_col + 1)]
    data = []
    for i in range(2, max_row + 1):
        row = []
        for j in range(1, max_col + 1):
            row.append(sheet_obj.cell(row=i, column=j).value)
        data.append(row)

    assert headers == expected_headers
    assert list(export_data[0].values()) in data
    row2 = list(export_data[1].values())
    # add string with space for collections column
    row2.append(" ")
    assert row2 in data

    shutil.rmtree(tmpdir)


@pytest.mark.skip(reason="CSV export rewritten, needs update")
@patch("saleor.csv.utils.export.BATCH_SIZE", 1)
def test_export_products_in_batches_for_csv(
    product_list, user_export_file, tmpdir, media_root,
):
    # given
    qs = Product.objects.all()
    export_info = {
        "fields": [ProductFieldEnum.NAME.value, ProductFieldEnum.VARIANT_SKU.value],
        "warehouses": [],
        "attributes": [],
    }
    file_name = "test.csv"
    export_fields = ["id", "name", "variants__sku"]
    expected_headers = ["id", "name", "variant sku"]

    table = etl.wrap([expected_headers])

    with NamedTemporaryFile() as temp_file:
        etl.tocsv(table, temp_file.name, delimiter=";")
        user_export_file.content_file.save(file_name, temp_file)

    assert user_export_file.content_file

    # when
    export_products_in_batches(
        qs,
        export_info,
        set(export_fields),
        export_fields,
        ";",
        user_export_file,
        FileTypes.CSV,
    )

    # then
    user_export_file.refresh_from_db()
    csv_file = user_export_file.content_file
    assert csv_file

    expected_data = []
    for product in qs.order_by("pk"):
        product_data = []
        product_data.append(str(product.pk))
        product_data.append(product.name)

        for variant in product.variants.all():
            product_data.append(str(variant.sku))
            expected_data.append(product_data)

    file_content = csv_file.read().decode().split("\r\n")

    # ensure headers are in file
    assert ";".join(expected_headers) in file_content

    for row in expected_data:
        assert ";".join(row) in file_content

    shutil.rmtree(tmpdir)


@pytest.mark.skip(reason="CSV export rewritten, needs update")
@patch("saleor.csv.utils.export.BATCH_SIZE", 1)
def test_export_products_in_batches_for_xlsx(
    product_list, user_export_file, tmpdir, media_root,
):
    # given
    qs = Product.objects.all()
    export_info = {
        "fields": [ProductFieldEnum.NAME.value, ProductFieldEnum.VARIANT_SKU.value],
        "warehouses": [],
        "attributes": [],
    }
    export_fields = ["id", "name", "variants__sku"]
    expected_headers = ["id", "name", "variant sku"]
    file_name = "test.xlsx"

    table = etl.wrap([expected_headers])

    with NamedTemporaryFile() as temp_file:
        etl.io.xlsx.toxlsx(table, temp_file.name)
        user_export_file.content_file.save(file_name, temp_file)

    assert user_export_file.content_file

    # when
    export_products_in_batches(
        qs,
        export_info,
        set(export_fields),
        export_fields,
        ";",
        user_export_file,
        FileTypes.XLSX,
    )

    # then
    user_export_file.refresh_from_db()
    assert user_export_file.content_file

    expected_data = []
    for product in qs.order_by("pk"):
        product_data = []
        product_data.append(product.pk)
        product_data.append(product.name)

        for variant in product.variants.all():
            product_data.append(variant.sku)
            expected_data.append(product_data)

    xlsx_file = user_export_file.content_file
    wb_obj = openpyxl.load_workbook(xlsx_file)

    sheet_obj = wb_obj.active
    max_col = sheet_obj.max_column
    max_row = sheet_obj.max_row
    headers = [sheet_obj.cell(row=1, column=i).value for i in range(1, max_col + 1)]
    data = []
    for i in range(2, max_row + 1):
        row = []
        for j in range(1, max_col + 1):
            row.append(sheet_obj.cell(row=i, column=j).value)
        data.append(row)

    assert headers == expected_headers
    for row in expected_data:
        assert row in data

    shutil.rmtree(tmpdir)
